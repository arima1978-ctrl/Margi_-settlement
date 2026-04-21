"""Sync プログラミング営業管理 sheet from a published Google Sheet CSV."""
from __future__ import annotations

import csv
import urllib.request
from pathlib import Path

from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from .sheet_replacer import column_letter_to_index


def fetch_csv(url_or_path: str) -> list[list[str]]:
    """Download a CSV from an HTTP(S) URL, or read from a local path.

    Returns a list of rows (each row is a list of cell strings).
    """
    if url_or_path.startswith("http://") or url_or_path.startswith("https://"):
        with urllib.request.urlopen(url_or_path, timeout=60) as resp:
            data = resp.read().decode("utf-8", errors="replace")
    else:
        data = Path(url_or_path).read_text(encoding="utf-8")

    reader = csv.reader(data.splitlines())
    return [row for row in reader]


def sync_sales_management(
    ws: Worksheet,
    csv_rows: list[list[str]],
    column_mapping: dict[int, str],
    csv_data_start_row: int,
    excel_data_start_row: int,
    excel_max_col_letter: str = "AG",
) -> int:
    """Replace rows in the sales management sheet using CSV data.

    Args:
        ws: Destination worksheet (プログラミング営業管理).
        csv_rows: Parsed CSV rows.
        column_mapping: {csv_col_index: excel_col_letter}
        csv_data_start_row: First CSV row index that contains shop data.
        excel_data_start_row: First Excel row to write.
        excel_max_col_letter: Rightmost column to clear (safety).

    Returns:
        Number of rows written.
    """
    # Step 1: clear existing data rows in the mapped column range
    excel_max_col = column_letter_to_index(excel_max_col_letter)
    for row in range(excel_data_start_row, ws.max_row + 1):
        for col in range(1, excel_max_col + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None

    # Step 2: write new rows
    mapped_items = list(column_mapping.items())  # [(csv_col, excel_col_letter), ...]
    written = 0
    for csv_row_idx in range(csv_data_start_row, len(csv_rows)):
        csv_row = csv_rows[csv_row_idx]
        # Skip rows with no 塾ID in the primary ID CSV column (col 2 by convention)
        if len(csv_row) <= 2 or not csv_row[2].strip():
            continue

        excel_row = excel_data_start_row + written
        for csv_col_idx, excel_col_letter in mapped_items:
            if csv_col_idx >= len(csv_row):
                continue
            raw = csv_row[csv_col_idx]
            val = _coerce(raw)
            excel_col_idx = column_letter_to_index(excel_col_letter)
            cell = ws.cell(row=excel_row, column=excel_col_idx)
            if isinstance(cell, MergedCell):
                continue
            cell.value = val
        written += 1
    return written


def _coerce(raw: str) -> object:
    """Convert CSV cell to int if it looks like a 塾ID or plain number, else str."""
    s = raw.strip()
    if not s:
        return None
    # Preserve #REF! and similar error tokens as strings
    if s.startswith("#"):
        return s
    # Try int
    if s.isdigit() or (s.startswith("-") and s[1:].isdigit()):
        try:
            return int(s)
        except ValueError:
            pass
    return s
