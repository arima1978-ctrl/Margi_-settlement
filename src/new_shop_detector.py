"""Detect new family IDs in source vs previous settlement.

A "new shop" means a family ID that has programming service activity
(appears in one of the ④ category sheets) but is not yet listed in the
previous settlement's 報告書 B column.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .sheet_replacer import column_letter_to_index


def get_family_ids_from_column(
    ws: Worksheet, column_letter: str, start_row: int = 2
) -> set[int]:
    """Read all integer family IDs from the given column."""
    col_idx = column_letter_to_index(column_letter)
    ids: set[int] = set()
    for row in range(start_row, ws.max_row + 1):
        val = ws.cell(row=row, column=col_idx).value
        if val is None:
            continue
        try:
            ids.add(int(val))
        except (TypeError, ValueError):
            continue
    return ids


def collect_active_family_ids(
    wb_src: Workbook,
    category_sheets: list[str],
    family_id_column: str = "C",
    data_start_row: int = 3,
) -> set[int]:
    """Collect family IDs that appear in any of the programming category sheets."""
    ids: set[int] = set()
    for sheet_name in category_sheets:
        if sheet_name not in wb_src.sheetnames:
            continue
        ws = wb_src[sheet_name]
        ids |= get_family_ids_from_column(ws, family_id_column, start_row=data_start_row)
    return ids


def detect_new_family_ids(
    wb_src: Workbook,
    category_sheets: list[str],
    dst_report_ws: Worksheet,
    dst_report_id_column: str,
    dst_report_start_row: int,
) -> list[int]:
    """Return family IDs with programming activity but not in dest 報告書 B column."""
    active_ids = collect_active_family_ids(wb_src, category_sheets)
    existing_ids = get_family_ids_from_column(
        dst_report_ws, dst_report_id_column, start_row=dst_report_start_row
    )
    new_ids = sorted(active_ids - existing_ids)
    return new_ids
