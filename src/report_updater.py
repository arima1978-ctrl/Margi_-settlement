"""Update 報告書 sheet: date cell, append new shops."""
from __future__ import annotations

from datetime import datetime

from openpyxl.worksheet.worksheet import Worksheet

from .sheet_replacer import column_letter_to_index


def update_month_cell(report_ws: Worksheet, cell_addr: str, month_date: datetime) -> None:
    """Set the target month date (e.g. 報告書!D1 = 2026-04-01)."""
    report_ws[cell_addr].value = month_date


def find_first_empty_row(ws: Worksheet, column_letter: str, start_row: int) -> int:
    """Return row right after the last non-empty cell in the column.

    This avoids hitting merged-cell 'dividers' that sit in the middle of the
    data range (e.g. 報告書 row 17 is a horizontal divider, so a naive
    first-empty scan would return 17 instead of the true append point).
    """
    col_idx = column_letter_to_index(column_letter)
    last_filled = start_row - 1
    for row in range(start_row, ws.max_row + 1):
        if ws.cell(row=row, column=col_idx).value is not None:
            last_filled = row
    return last_filled + 1


def append_family_ids_to_report(
    report_ws: Worksheet,
    family_ids: list[int],
    b_column: str = "B",
    al_column: str | None = "AL",
    data_start_row: int = 11,
) -> list[int]:
    """Append new family IDs to B column (and optionally AL) of 報告書.

    Returns the rows used. Skips IDs already present in the B column.
    Raises if target cells are merged. Pass ``al_column=None`` for reports
    that do not mirror the family ID to AL (e.g. 将棋/文理/速読).
    """
    from openpyxl.cell.cell import MergedCell

    # Collect existing IDs to avoid duplicates
    b_idx = column_letter_to_index(b_column)
    al_idx = column_letter_to_index(al_column) if al_column else None
    existing = set()
    for row in range(data_start_row, report_ws.max_row + 1):
        val = report_ws.cell(row=row, column=b_idx).value
        if val is not None:
            try:
                existing.add(int(val))
            except (TypeError, ValueError):
                pass

    first_empty = find_first_empty_row(report_ws, b_column, data_start_row)
    rows_used: list[int] = []
    cursor = first_empty
    for fid in family_ids:
        if fid in existing:
            continue
        b_cell = report_ws.cell(row=cursor, column=b_idx)
        if isinstance(b_cell, MergedCell):
            raise RuntimeError(
                f"Cannot append to row {cursor}: B cell is part of a merged range. "
                "Check 報告書 sheet's merged cells and adjust data_start_row."
            )
        if al_idx is not None:
            al_cell = report_ws.cell(row=cursor, column=al_idx)
            if isinstance(al_cell, MergedCell):
                raise RuntimeError(
                    f"Cannot append to row {cursor}: AL cell is part of a merged range. "
                    "Check 報告書 sheet's merged cells or set report_al_column: null."
                )
            al_cell.value = f"=B{cursor}"
        b_cell.value = fid
        rows_used.append(cursor)
        existing.add(fid)
        cursor += 1
    return rows_used
