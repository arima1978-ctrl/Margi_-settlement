"""Eduplus monthly aggregation processor.

What it does
------------
Operates **in place** on the source ``.xlsm`` backup file. For each monthly
run it:

1. Aggregates each eduplus source sheet's L/M (= C/H) into unique
   (family_id, summed_amount) pairs and writes the result back into
   columns N and O, starting at row 5.
2. Populates ``学書マージン清算書用シート``:
     - Columns A/B (row 7+): concatenated N/O pairs from the 4 source
       sheets — family IDs may appear multiple times across sheets.
     - Columns D/E (row 7+): deduplicated + summed across sheets.
     - Column G: ``=COUNTIF(マージン計算用!E:E, D{row})``
     - Column H: ``=IF(G{row}=1,"","新規")``
3. Appends family IDs that are missing from ``マージン計算用!E:E`` to that
   sheet's E column (first empty row).

Idempotency
-----------
Safe to re-run. Steps 1 and 2 clear their targets before writing. Step 3
only appends IDs that are not already in the destination column, so
successive runs become no-ops once all 新規 have been absorbed.

Macros (VBA) in the ``.xlsm`` are preserved via ``keep_vba=True``.
"""
from __future__ import annotations

import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


# Each source sheet starts its data at a different row: ⑤ has two header
# rows (r3/r4) so data starts at r5, while ⑥/⑥-2/⑥-3 only have one header
# (r3) and real data begins at r4. Missing this nuance silently dropped
# three families (100333, 101301, 101725) on the first implementation.
EDUPLUS_SOURCE_SHEETS: tuple[tuple[str, int], ...] = (
    ("⑤Edu　ID利用料",          5),
    ("⑥Edu　基本料金DLデータ",  4),
    ("⑥-2 Edu初期費用　",       4),
    ("⑥-3 Edu解約塾",            4),
)

SHUGO_SHEET = "学書マージン清算書用シート"
MARGIN_KEISAN_SHEET = "マージン計算用"

# Each eduplus source sheet uses the same column layout.
COL_C = 3    # 家族ID
COL_H = 8    # 料金
COL_N = 14   # 学書ID (aggregation output)
COL_O = 15   # 学書料金 (aggregation output)

SHUGO_DATA_START = 7        # 学書マージン清算書用シート データ行開始
MARGIN_KEISAN_ID_COL = 5    # E列 = 家族ID


@dataclass
class SheetResult:
    name: str
    found: bool = True
    unique_families: int = 0
    rows_scanned: int = 0


@dataclass
class EduplusResult:
    source_path: Path
    sheets: list[SheetResult] = field(default_factory=list)
    total_unique_across_sheets: int = 0  # = rows written to A/B in 学書マージン清算書
    total_deduped_across_sheets: int = 0  # = rows written to D/E
    new_family_ids: list[int] = field(default_factory=list)
    margin_keisan_append_from_row: int | None = None


def aggregate_source_sheet(ws: Worksheet, data_start_row: int = 5
                           ) -> tuple[list[tuple[int, float]], int]:
    """Aggregate C/H into unique (family_id, summed_amount) written to N/O.

    Returns (pairs_in_order, rows_scanned). Also clears the old N/O region
    before writing so stale data from previous runs is removed.
    """
    totals: dict[int, float] = {}
    order: list[int] = []
    rows_scanned = 0

    for row in range(data_start_row, ws.max_row + 1):
        c = ws.cell(row=row, column=COL_C).value
        if c is None or c == "":
            continue
        try:
            family_id = int(c)
        except (TypeError, ValueError):
            continue
        rows_scanned += 1

        h = ws.cell(row=row, column=COL_H).value
        amount = float(h) if isinstance(h, (int, float)) else 0.0

        if family_id not in totals:
            totals[family_id] = 0.0
            order.append(family_id)
        totals[family_id] += amount

    # Clear any previous N/O from data_start_row down to max_row
    for row in range(data_start_row, ws.max_row + 1):
        ws.cell(row=row, column=COL_N).value = None
        ws.cell(row=row, column=COL_O).value = None

    pairs: list[tuple[int, float]] = [(fid, totals[fid]) for fid in order]
    for i, (fid, amount) in enumerate(pairs):
        target_row = data_start_row + i
        ws.cell(row=target_row, column=COL_N).value = fid
        ws.cell(row=target_row, column=COL_O).value = _as_int_if_whole(amount)

    return pairs, rows_scanned


def _as_int_if_whole(v: float) -> int | float:
    """Preserve the visual of 'no decimals' when the amount is a whole number."""
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def write_shugo_sheet(ws: Worksheet, all_pairs: list[tuple[int, float]]) -> tuple[int, list[tuple[int, float]]]:
    """Populate A/B with concatenated pairs and D/E with deduplicated pairs.

    Returns (rows_written_to_AB, deduped_pairs_in_D_E). The existing totals
    formulas in E4 (=SUM(E7:E1391)) and E5 (=E4*1.08) are preserved so
    cached totals shown at the top of the sheet stay in sync after a
    recalc. Only the data rows (r7 onward) are cleared before writing.
    """
    # Clear A, B, D, E, G, H from SHUGO_DATA_START down to max_row.
    # Rows above SHUGO_DATA_START contain title/header and the SUM totals
    # (E4/E5), which must be left untouched.
    for row in range(SHUGO_DATA_START, ws.max_row + 1):
        for col in (1, 2, 4, 5, 7, 8):
            ws.cell(row=row, column=col).value = None

    # Write A/B (concatenated, duplicates across sheets kept)
    for i, (fid, amount) in enumerate(all_pairs):
        target = SHUGO_DATA_START + i
        ws.cell(row=target, column=1).value = fid
        ws.cell(row=target, column=2).value = _as_int_if_whole(amount)

    # Dedup + sum across sheets
    dedup_totals: dict[int, float] = {}
    dedup_order: list[int] = []
    for fid, amount in all_pairs:
        if fid not in dedup_totals:
            dedup_totals[fid] = 0.0
            dedup_order.append(fid)
        dedup_totals[fid] += amount

    deduped: list[tuple[int, float]] = [(fid, dedup_totals[fid]) for fid in dedup_order]

    for i, (fid, amount) in enumerate(deduped):
        target = SHUGO_DATA_START + i
        ws.cell(row=target, column=4).value = fid
        ws.cell(row=target, column=5).value = _as_int_if_whole(amount)
        # G: COUNTIF against マージン計算用 E column
        ws.cell(row=target, column=7).value = (
            f"=COUNTIF(マージン計算用!E:E,学書マージン清算書用シート!D{target})"
        )
        # H: IF(G=1, "", "新規") — 新規 appears when the family is NOT yet in マージン計算用
        ws.cell(row=target, column=8).value = f'=IF(G{target}=1,"","新規")'

    return len(all_pairs), deduped


def append_new_ids_to_margin_keisan(ws: Worksheet, deduped: list[tuple[int, float]]) -> tuple[list[int], int | None]:
    """Append family IDs missing from マージン計算用!E to the first empty row.

    Returns (new_ids, start_row). ``start_row`` is None when there is nothing
    to add.
    """
    existing: set[int] = set()
    last_filled = 0
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=MARGIN_KEISAN_ID_COL).value
        if v is None or v == "":
            continue
        try:
            fid = int(v)
        except (TypeError, ValueError):
            # Non-numeric entries (headers like '家族ID') still count as
            # 'the column was touched up to this row'.
            last_filled = row
            continue
        existing.add(fid)
        last_filled = row

    new_ids: list[int] = [fid for fid, _ in deduped if fid not in existing]
    if not new_ids:
        return [], None

    start_row = last_filled + 1
    for i, fid in enumerate(new_ids):
        ws.cell(row=start_row + i, column=MARGIN_KEISAN_ID_COL).value = fid

    return new_ids, start_row


def process_eduplus(source_path: str | Path, *, backup: bool = True) -> EduplusResult:
    """Run the full eduplus monthly processing on the source .xlsm in place.

    With ``backup=True`` (default), a ``.bak`` copy is written next to the
    source before any modification.
    """
    path = Path(source_path)
    if not path.exists():
        raise FileNotFoundError(f"source .xlsm not found: {path}")
    if path.suffix.lower() != ".xlsm":
        raise ValueError(f"source must be .xlsm, got {path.suffix}")

    if backup:
        shutil.copy2(path, path.with_suffix(".xlsm.bak"))

    result = EduplusResult(source_path=path)

    # keep_vba=True: the business workbook contains macros that must survive.
    wb = openpyxl.load_workbook(path, keep_vba=True)

    all_pairs: list[tuple[int, float]] = []
    for sheet_name, start_row in EDUPLUS_SOURCE_SHEETS:
        if sheet_name not in wb.sheetnames:
            result.sheets.append(SheetResult(name=sheet_name, found=False))
            continue
        pairs, scanned = aggregate_source_sheet(wb[sheet_name], data_start_row=start_row)
        all_pairs.extend(pairs)
        result.sheets.append(SheetResult(
            name=sheet_name, found=True,
            unique_families=len(pairs), rows_scanned=scanned,
        ))

    if SHUGO_SHEET in wb.sheetnames:
        written_ab, deduped = write_shugo_sheet(wb[SHUGO_SHEET], all_pairs)
        result.total_unique_across_sheets = written_ab
        result.total_deduped_across_sheets = len(deduped)
    else:
        deduped = []

    if MARGIN_KEISAN_SHEET in wb.sheetnames and deduped:
        new_ids, start_row = append_new_ids_to_margin_keisan(wb[MARGIN_KEISAN_SHEET], deduped)
        result.new_family_ids = new_ids
        result.margin_keisan_append_from_row = start_row

    # Ensure Excel fully recalculates formulas (G/H and others) on next open.
    wb.calculation.fullCalcOnLoad = True

    wb.save(path)
    wb.close()
    return result


def format_summary(result: EduplusResult) -> str:
    """Human-friendly summary (used by CLI and Telegram notifier)."""
    lines: list[str] = [f"【eduplus 集計】{result.source_path.name}"]
    for s in result.sheets:
        if s.found:
            lines.append(f"  {s.name}: {s.rows_scanned}行 → unique {s.unique_families}塾")
        else:
            lines.append(f"  {s.name}: NOT FOUND")
    lines.append(f"学書マージン清算書用シート: A/B={result.total_unique_across_sheets}行, "
                 f"D/E={result.total_deduped_across_sheets}行")
    if result.new_family_ids:
        lines.append(f"マージン計算用 E列に新規追加: {len(result.new_family_ids)}塾 "
                     f"(r{result.margin_keisan_append_from_row}〜)")
        lines.append("  新規家族ID: " + ", ".join(str(i) for i in result.new_family_ids[:20]))
        if len(result.new_family_ids) > 20:
            lines.append(f"  ... 他 {len(result.new_family_ids) - 20} 件")
    else:
        lines.append("マージン計算用: 新規追加なし")
    return "\n".join(lines)
