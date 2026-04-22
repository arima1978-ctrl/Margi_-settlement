"""Detect contracted shops from a master sheet to auto-populate 報告書."""
from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from .sheet_replacer import column_letter_to_index


def detect_contract_shops(
    master_ws: Worksheet,
    id_column: str,
    status_column: str,
    status_value: str,
    data_start_row: int,
) -> list[int]:
    """Return family IDs from master sheet whose status column matches status_value.

    Preserves master-sheet order so downstream append writes in deterministic
    order.
    """
    id_col_idx = column_letter_to_index(id_column)
    status_col_idx = column_letter_to_index(status_column)

    ids: list[int] = []
    for row in range(data_start_row, master_ws.max_row + 1):
        status = master_ws.cell(row=row, column=status_col_idx).value
        if status != status_value:
            continue
        raw = master_ws.cell(row=row, column=id_col_idx).value
        if raw is None:
            continue
        try:
            ids.append(int(raw))
        except (TypeError, ValueError):
            continue
    return ids


def filter_missing_from_report(
    ids: list[int],
    report_ws: Worksheet,
    report_id_column: str,
    report_data_start_row: int,
) -> list[int]:
    """Filter ids down to those not already present in report_ws's id column."""
    id_col_idx = column_letter_to_index(report_id_column)
    existing: set[int] = set()
    for row in range(report_data_start_row, report_ws.max_row + 1):
        val = report_ws.cell(row=row, column=id_col_idx).value
        if val is None:
            continue
        try:
            existing.add(int(val))
        except (TypeError, ValueError):
            continue
    return [fid for fid in ids if fid not in existing]
