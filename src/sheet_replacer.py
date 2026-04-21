"""Replace data in settlement category sheets from source .xlsm."""
from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet


def column_letter_to_index(letter: str) -> int:
    """'A' -> 1, 'B' -> 2, 'AA' -> 27."""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def clear_data_rows(ws: Worksheet, data_start_row: int) -> None:
    """Remove all cells from data_start_row to max_row, preserving headers."""
    if ws.max_row < data_start_row:
        return
    for row in range(data_start_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None


def copy_data_rows(
    src_ws: Worksheet,
    dst_ws: Worksheet,
    data_start_row: int,
    columns_to_copy: list[str] | None = None,
) -> int:
    """Copy data rows from src to dst (values only, no formulas).

    Returns the number of data rows copied.
    """
    if columns_to_copy is None:
        # Copy all columns from src
        max_col = src_ws.max_column
        col_indices = list(range(1, max_col + 1))
    else:
        col_indices = [column_letter_to_index(c) for c in columns_to_copy]

    rows_copied = 0
    for src_row_idx in range(data_start_row, src_ws.max_row + 1):
        # Read source row values (resolve formulas by taking cached value)
        has_data = False
        values = {}
        for col_idx in col_indices:
            cell = src_ws.cell(row=src_row_idx, column=col_idx)
            val = cell.value
            # If it's a formula string, we cannot easily resolve without
            # data_only workbook. Caller should pass a data_only source.
            values[col_idx] = val
            if val is not None and val != "":
                has_data = True

        if not has_data:
            # Preserve blank rows within data range by still writing None,
            # but stop once we see a fully empty row (optimization).
            continue

        dst_row_idx = data_start_row + rows_copied
        for col_idx, val in values.items():
            dst_ws.cell(row=dst_row_idx, column=col_idx).value = val
        rows_copied += 1

    return rows_copied


def apply_helper_columns(
    ws: Worksheet,
    data_start_row: int,
    num_rows: int,
    helper_columns: dict[str, str],
) -> None:
    """Apply helper columns (L/M/O/P) used by 報告書 VLOOKUPs.

    helper_columns is a dict like:
        {"L": "=C{row}", "M": "=H{row}", "O": "C", "P": "H"}
    - Formula spec starts with "=" and uses "{row}" placeholder
    - Raw spec is a column letter, meaning "copy value from that column"
    """
    if not helper_columns:
        return

    for helper_col, spec in helper_columns.items():
        helper_col_idx = column_letter_to_index(helper_col)

        if spec.startswith("="):
            # Formula: write formula string with row substituted
            for row in range(data_start_row, data_start_row + num_rows):
                # Only write if the source cell has data (avoid formulas on empty rows)
                ws.cell(row=row, column=helper_col_idx).value = spec.format(row=row)
        else:
            # Raw copy: spec is source column letter (e.g. "C")
            src_col_idx = column_letter_to_index(spec)
            for row in range(data_start_row, data_start_row + num_rows):
                src_val = ws.cell(row=row, column=src_col_idx).value
                ws.cell(row=row, column=helper_col_idx).value = src_val


def clear_column_range(ws: Worksheet, col_letter: str, from_row: int) -> None:
    """Clear a single column from from_row to max_row."""
    col_idx = column_letter_to_index(col_letter)
    for row in range(from_row, ws.max_row + 1):
        ws.cell(row=row, column=col_idx).value = None


def apply_aggregates(
    ws: Worksheet,
    data_start_row: int,
    num_rows: int,
    aggregates: list[dict],
) -> None:
    """Write aggregated columns (e.g. O=unique 家族ID, P=sum of 料金 per ID).

    Each aggregate spec:
        dst_col_id:  destination column letter for unique IDs (e.g. "O")
        dst_col_sum: destination column letter for sum values (e.g. "P")
        src_col_id:  source column letter to group by (e.g. "C")
        src_col_sum: source column letter to sum (e.g. "H")

    Walks the data region and for each unique ID in src_col_id, writes one row
    (id, sum) into (dst_col_id, dst_col_sum) starting at data_start_row.
    """
    if not aggregates:
        return

    for agg in aggregates:
        src_id_idx = column_letter_to_index(agg["src_col_id"])
        src_sum_idx = column_letter_to_index(agg["src_col_sum"])
        dst_id_idx = column_letter_to_index(agg["dst_col_id"])
        dst_sum_idx = column_letter_to_index(agg["dst_col_sum"])

        # Clear destination aggregate columns across the full data range to
        # remove any source-side garbage (COUNTIF/"OK" etc.) that leaked in
        # via the raw row copy.
        for row in range(data_start_row, ws.max_row + 1):
            ws.cell(row=row, column=dst_id_idx).value = None
            ws.cell(row=row, column=dst_sum_idx).value = None

        totals: dict[object, float] = {}
        order: list[object] = []
        for row in range(data_start_row, data_start_row + num_rows):
            src_id = ws.cell(row=row, column=src_id_idx).value
            if src_id is None or src_id == "":
                continue
            src_sum = ws.cell(row=row, column=src_sum_idx).value
            if not isinstance(src_sum, (int, float)):
                continue
            if src_id not in totals:
                totals[src_id] = 0
                order.append(src_id)
            totals[src_id] += src_sum

        for i, key in enumerate(order):
            target_row = data_start_row + i
            ws.cell(row=target_row, column=dst_id_idx).value = key
            total = totals[key]
            # Preserve int-ness when sum has no fractional part
            if isinstance(total, float) and total.is_integer():
                total = int(total)
            ws.cell(row=target_row, column=dst_sum_idx).value = total


def replace_sheet_data(
    src_ws: Worksheet,
    dst_ws: Worksheet,
    data_start_row: int,
    helper_columns: dict[str, str],
    aggregates: list[dict] | None = None,
) -> int:
    """Replace destination sheet's data rows with source's data, values-only.

    Returns the number of rows copied.
    """
    # 1. Clear destination data rows
    clear_data_rows(dst_ws, data_start_row)

    # 2. Copy data from source (values only)
    rows_copied = copy_data_rows(src_ws, dst_ws, data_start_row)

    # 3. Apply helper columns
    apply_helper_columns(dst_ws, data_start_row, rows_copied, helper_columns)

    # 4. Apply aggregates (deduplication + sum) — used by 報告書 VLOOKUPs
    apply_aggregates(dst_ws, data_start_row, rows_copied, aggregates or [])

    return rows_copied
