"""eteacher売上管理表 の月次更新ツール.

Workflow
--------
1. マージン計算用 の各行から (塾名, 家族ID, T=売上) を算出
   - T = ROUNDDOWN((初期費用 + 基本料金 + 利用料金 + 解約分) * 1.1, 0)
   - 各成分は eduplus_processor が書き込んだ 4 シートの N/O を Python で
     VLOOKUP 相当の処理で取得する
2. 3月.xls テンプレ を .xlsx にコピー/変換
3. openpyxl で D列の前に「家族ID」列を新規挿入(既存データ 1列ずつ右シフト)
4. 各データ行で塾名照合 → 新 D 列に家族ID、新 AO 列に売上を書く
5. 新 r7 に「入金日」「売上額」ヘッダーを追加
"""
from __future__ import annotations

import math
import re
import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl


# ---- lookup from processed xlsm -----------------------------------------

_SHEET_START_ROWS = {
    "⑤Edu　ID利用料":          5,
    "⑥Edu　基本料金DLデータ":  4,
    "⑥-2 Edu初期費用　":       4,
    "⑥-3 Edu解約塾":            4,
}


def _build_family_amount_lookup(wb: openpyxl.Workbook, sheet_name: str) -> dict[int, float]:
    """Read N (family_id) / O (amount) columns from a source sheet.

    Our eduplus_processor writes unique (family_id, sum_amount) pairs
    starting at that sheet's data_start_row, so no VLOOKUP is needed.
    """
    if sheet_name not in wb.sheetnames:
        return {}
    start_row = _SHEET_START_ROWS.get(sheet_name, 4)
    ws = wb[sheet_name]
    out: dict[int, float] = {}
    for row in range(start_row, ws.max_row + 1):
        key = ws.cell(row=row, column=14).value  # N
        val = ws.cell(row=row, column=15).value  # O
        if key is None or key == "":
            continue
        try:
            fid = int(key)
        except (TypeError, ValueError):
            continue
        if isinstance(val, (int, float)):
            out[fid] = float(val)
        else:
            out[fid] = 0.0
    return out


def _roundown_toward_zero(x: float) -> int:
    """Excel-style ROUNDDOWN(x, 0) — truncate toward zero."""
    return math.trunc(x)


@dataclass
class ShopSales:
    family_id: int
    sales: int                 # ROUNDDOWN(base * 1.1, 0)
    base: float                # 税抜 = O + P + Q + R
    name: str = ""             # 塾名 (from 保護者情報 B列)
    tel: str = ""              # 電話番号 (保護者情報 H列)
    addr: str = ""             # 住所 (保護者情報 G列)
    rep: str = ""              # 代表者 (らくらく ユーザー基本情報 G列)


def _build_hogosha_info_lookup(wb: openpyxl.Workbook) -> dict[int, tuple[str, str, str]]:
    """Build 家族ID -> (塾名, 電話番号, 住所) from 保護者情報DL貼付⑩AKへ.

    Column layout:  A=家族ID, B=塾名(氏名), G=住所, H=電話番号.

    Used both as fallback when ``マージン計算用`` D column is a formula (no
    cached value) and as the source for cross-reference matching (TEL/住所)
    when 塾名 differs between eteacher and the source workbook.
    """
    sheet_name = "保護者情報DL貼付⑩AKへ"
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    out: dict[int, tuple[str, str, str]] = {}
    for row in range(2, ws.max_row + 1):
        raw_fid = ws.cell(row=row, column=1).value
        name = ws.cell(row=row, column=2).value
        addr = ws.cell(row=row, column=7).value
        tel = ws.cell(row=row, column=8).value
        try:
            fid = int(raw_fid)
        except (TypeError, ValueError):
            continue
        out[fid] = (
            str(name).strip() if isinstance(name, str) else "",
            str(tel).strip() if tel is not None else "",
            str(addr).strip() if isinstance(addr, str) else "",
        )
    return out


def normalize_tel(s: Any) -> str:
    """Strip all non-digit characters so TELs like '0265-52-4119' and
    '026552 4119' compare equal."""
    if not s:
        return ""
    return re.sub(r"\D", "", str(s))


def normalize_addr(s: Any) -> str:
    """Strip spaces (全角/半角) and newlines for 住所 comparison.

    eteacher 側は短縮住所 (都道府県+市区町村) が多く、保護者情報 側は完全
    住所 が多いので、呼び出し側で prefix 一致を使う前提の正規化にとどめる。
    """
    if not s:
        return ""
    return re.sub(r"[\s　]", "", str(s))


# Common corporate prefixes to strip when computing a canonical shop name.
_CORPORATE_PREFIXES = (
    "株式会社", "(株)", "(株)", "（株）",
    "有限会社", "(有)", "(有)", "（有）",
    "合同会社", "(合)", "(合)", "（合）",
    "学校法人",
    "特定非営利活動法人",
)


def normalize_shop_name(s: Any) -> str:
    """Strip 法人格 prefix/suffix and spaces for fuzzy shop-name matching.

    e.g. "(株)アール塾" and "アール塾" become the same canonical form.
    """
    if not s:
        return ""
    n = str(s).strip()
    for token in _CORPORATE_PREFIXES:
        n = n.replace(token, "")
    n = re.sub(r"[\s　]", "", n)
    return n


def _build_rakuraku_rep_lookup(wb: openpyxl.Workbook) -> dict[int, str]:
    """Build 家族ID -> 代表者 from らくらく ユーザー基本情報 G列."""
    sheet_name = "らくらく　ユーザー基本情報貼り付ける"
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    out: dict[int, str] = {}
    for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        raw_fid = row[0]
        g = row[6] if len(row) > 6 else None
        if raw_fid is None or not g:
            continue
        try:
            fid = int(raw_fid)
        except (TypeError, ValueError):
            continue
        if isinstance(g, str) and g.strip():
            # Collapse multiple spaces to keep display tidy
            out[fid] = re.sub(r"\s+", " ", g.strip())
    return out


def compute_shops_with_sales(xlsm_path: str | Path) -> list[ShopSales]:
    """Return a list of ShopSales (one per unique family_id).

    Each entry carries the computed monthly sales AND contact info
    (塾名/電話番号/住所) from 保護者情報DL貼付⑩AKへ, enabling multi-field
    matching downstream.
    """
    wb = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)

    shoki = _build_family_amount_lookup(wb, "⑥-2 Edu初期費用　")
    kihon = _build_family_amount_lookup(wb, "⑥Edu　基本料金DLデータ")
    riyo = _build_family_amount_lookup(wb, "⑤Edu　ID利用料")
    kaiyaku = _build_family_amount_lookup(wb, "⑥-3 Edu解約塾")
    hogosha = _build_hogosha_info_lookup(wb)
    rep_lookup = _build_rakuraku_rep_lookup(wb)

    ws = wb["マージン計算用"]
    seen: set[int] = set()
    out: list[ShopSales] = []
    for row in range(12, ws.max_row + 1):
        raw_fid = ws.cell(row=row, column=5).value
        try:
            fid = int(raw_fid) if raw_fid is not None else None
        except (TypeError, ValueError):
            fid = None
        if fid is None or fid in seen:
            continue
        seen.add(fid)

        # Prefer cached value from マージン計算用 D, fall back to 保護者情報.
        name_raw = ws.cell(row=row, column=4).value
        if isinstance(name_raw, str) and name_raw.strip() and not name_raw.startswith("="):
            name = name_raw.strip()
        else:
            name = hogosha.get(fid, ("", "", ""))[0]

        tel = hogosha.get(fid, ("", "", ""))[1]
        addr = hogosha.get(fid, ("", "", ""))[2]
        rep = rep_lookup.get(fid, "")

        base = (
            shoki.get(fid, 0.0)
            + kihon.get(fid, 0.0)
            + riyo.get(fid, 0.0)
            + kaiyaku.get(fid, 0.0)
        )
        t = _roundown_toward_zero(base * 1.1)
        out.append(ShopSales(family_id=fid, sales=t, base=base,
                             name=name or "", tel=tel, addr=addr, rep=rep))

    wb.close()
    return out


def compute_sales_by_shop(xlsm_path: str | Path) -> dict[str, ShopSales]:
    """Backwards-compat wrapper: name -> ShopSales (only those with a name)."""
    return {s.name: s for s in compute_shops_with_sales(xlsm_path) if s.name}


# ---- .xls -> .xlsx conversion via Excel COM -----------------------------

def convert_xls_to_xlsx(src_xls: str | Path, dst_xlsx: str | Path) -> None:
    """Convert .xls -> .xlsx using Excel via win32com.

    Requires Microsoft Excel installed on the machine. This is run on the
    user's Windows PC, not on the Linux cron server.
    """
    import win32com.client  # type: ignore[import]

    src_abs = str(Path(src_xls).resolve())
    dst_abs = str(Path(dst_xlsx).resolve())
    dst_parent = Path(dst_xlsx).resolve().parent
    dst_parent.mkdir(parents=True, exist_ok=True)

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(src_abs)
        # FileFormat 51 = xlOpenXMLWorkbook (.xlsx)
        wb.SaveAs(dst_abs, FileFormat=51)
        wb.Close(SaveChanges=False)
    finally:
        excel.Quit()


# ---- eteacher .xlsx modification ----------------------------------------

@dataclass
class EteacherUpdateResult:
    template_path: Path
    output_path: Path
    matched: list[tuple[str, int, int, str]] = field(default_factory=list)   # (eteacher塾名, 家族ID, 売上, match_method)
    unmatched_in_eteacher: list[str] = field(default_factory=list)           # eteacher にあって source 側で解決できなかった塾
    unmatched_in_source: list[str] = field(default_factory=list)             # source にあるが eteacher 側に対応する行が無い塾
    match_method_counts: dict[str, int] = field(default_factory=dict)         # name / tel / addr の内訳


# Column indices (1-based) in the 3月 eteacher template (before shift).
_COL_SHOP_NAME = 4       # D = 塾名
_COL_ADDR = 8            # H = 住所 (都道府県+市区町村 の短縮形が多い)
_COL_TEL = 9             # I = 電話番号
_COL_HEADER_ROW = 6      # r6 has '塾名' header
_COL_MONTH_HEADER_ROW = 7  # r7 has 入金日/売上額 pattern


def _match_shop(name: str, tel: str, addr: str,
                shops_by_name: dict[str, ShopSales],
                shops_by_tel: dict[str, ShopSales],
                shops_by_canon_name: dict[str, ShopSales],
                shops: list[ShopSales]) -> tuple[ShopSales | None, str]:
    """Try 4-stage matching and return (shop, method).

    1. 塾名 exact         (method='name')
    2. TEL 正規化一致     (method='tel')
    3. 住所 prefix 一致   (method='addr')
    4. 塾名 部分一致       (method='name_partial')
       - canonical form: strip 法人格 prefix + spaces, then compare
       - also tries substring containment (eteacher ⊂ source or逆)
    """
    key = (name or "").strip()
    if key and key in shops_by_name:
        return shops_by_name[key], "name"

    t_norm = normalize_tel(tel)
    if t_norm and t_norm in shops_by_tel:
        return shops_by_tel[t_norm], "tel"

    a_norm = normalize_addr(addr)
    if a_norm:
        for s in shops:
            sa = normalize_addr(s.addr)
            if not sa:
                continue
            if sa.startswith(a_norm) or a_norm.startswith(sa):
                return s, "addr"

    # Fuzzy: normalized canonical name equality
    n_canon = normalize_shop_name(key)
    if n_canon and n_canon in shops_by_canon_name:
        return shops_by_canon_name[n_canon], "name_partial"

    # Fuzzy: substring containment on canonical form, min 4 chars to avoid
    # '塾' alone colliding everywhere.
    if n_canon and len(n_canon) >= 4:
        for s in shops:
            s_canon = normalize_shop_name(s.name)
            if not s_canon or len(s_canon) < 4:
                continue
            if n_canon in s_canon or s_canon in n_canon:
                return s, "name_partial"

    return None, ""


def update_eteacher(
    xlsx_path: str | Path,
    shops: list[ShopSales],
    insert_family_id_col: bool = True,
) -> EteacherUpdateResult:
    """Insert 家族ID column before D, then write this month's 売上 to AO.

    Modifies ``xlsx_path`` in place — the caller is responsible for making
    a copy of the template first (via ``convert_xls_to_xlsx``).

    Matching is tried in order: 塾名 exact → TEL 正規化一致 → 住所 prefix 一致.
    """
    output_xlsx = Path(xlsx_path)

    wb = openpyxl.load_workbook(output_xlsx)
    ws = wb.active  # Sheet2 in the template

    result = EteacherUpdateResult(template_path=output_xlsx, output_path=output_xlsx)

    # Build lookup dicts for fast matching
    shops_by_name = {s.name.strip(): s for s in shops if s.name}
    shops_by_tel: dict[str, ShopSales] = {}
    for s in shops:
        t = normalize_tel(s.tel)
        if t:
            shops_by_tel.setdefault(t, s)
    shops_by_canon_name: dict[str, ShopSales] = {}
    for s in shops:
        c = normalize_shop_name(s.name)
        if c:
            shops_by_canon_name.setdefault(c, s)

    # Snapshot each eteacher row's 塾名/TEL/住所 BEFORE inserting the column.
    # The snapshot lets matching operate on stable column indices.
    row_info: dict[int, tuple[str, str, str]] = {}
    max_row = ws.max_row
    for row in range(8, max_row + 1):
        name = ws.cell(row=row, column=_COL_SHOP_NAME).value
        addr = ws.cell(row=row, column=_COL_ADDR).value
        tel = ws.cell(row=row, column=_COL_TEL).value
        if isinstance(name, str) and name.strip():
            row_info[row] = (name.strip(),
                             str(tel).strip() if tel is not None else "",
                             str(addr).strip() if isinstance(addr, str) else "")

    if insert_family_id_col:
        ws.insert_cols(_COL_SHOP_NAME, 1)
        family_id_col = _COL_SHOP_NAME
        ws.cell(row=_COL_HEADER_ROW, column=family_id_col).value = "家族ID"
    else:
        family_id_col = None

    # New month columns: shifted right by 1 if we inserted the 家族ID col
    if insert_family_id_col:
        new_month_date_col = 40  # AN
        new_month_sales_col = 41  # AO
    else:
        new_month_date_col = 39  # AM
        new_month_sales_col = 40  # AN

    ws.cell(row=_COL_MONTH_HEADER_ROW, column=new_month_date_col).value = "入金日"
    ws.cell(row=_COL_MONTH_HEADER_ROW, column=new_month_sales_col).value = "売上額"

    # Reference columns: show the matched shop's source-side data so the
    # operator can eyeball whether the match is correct.
    ref_addr_col = new_month_sales_col + 1      # AP (or AO if no insert)
    ref_tel_col = new_month_sales_col + 2       # AQ
    ref_rep_col = new_month_sales_col + 3       # AR
    ref_method_col = new_month_sales_col + 4    # AS
    ws.cell(row=_COL_HEADER_ROW, column=ref_addr_col).value = "参照住所"
    ws.cell(row=_COL_HEADER_ROW, column=ref_tel_col).value = "参照TEL"
    ws.cell(row=_COL_HEADER_ROW, column=ref_rep_col).value = "参照代表者"
    ws.cell(row=_COL_HEADER_ROW, column=ref_method_col).value = "照合方法"

    matched_family_ids: set[int] = set()
    for row_idx, (name, tel, addr) in row_info.items():
        shop, method = _match_shop(
            name, tel, addr,
            shops_by_name, shops_by_tel, shops_by_canon_name, shops,
        )
        if shop is None:
            result.unmatched_in_eteacher.append(name)
            continue

        if family_id_col is not None:
            ws.cell(row=row_idx, column=family_id_col).value = shop.family_id
        if shop.sales != 0:
            ws.cell(row=row_idx, column=new_month_sales_col).value = shop.sales

        # Reference info (always populated for matched rows)
        ws.cell(row=row_idx, column=ref_addr_col).value = shop.addr or None
        ws.cell(row=row_idx, column=ref_tel_col).value = shop.tel or None
        ws.cell(row=row_idx, column=ref_rep_col).value = shop.rep or None
        ws.cell(row=row_idx, column=ref_method_col).value = method

        matched_family_ids.add(shop.family_id)
        result.matched.append((name, shop.family_id, shop.sales, method))
        result.match_method_counts[method] = result.match_method_counts.get(method, 0) + 1

    # Shops present in source but not found in eteacher
    result.unmatched_in_source = sorted(
        f"{s.name or '(無名)'} (ID={s.family_id})"
        for s in shops if s.family_id not in matched_family_ids
    )

    wb.save(output_xlsx)
    wb.close()
    return result


def format_eteacher_summary(result: EteacherUpdateResult, detail_preview: int = 5) -> str:
    """Short summary used in CLI stdout and Telegram. See ``write_unmatched_report``
    for the full list intended for operator review."""
    method_parts = ", ".join(
        f"{k}={v}" for k, v in sorted(result.match_method_counts.items())
    ) or "(なし)"
    lines = [
        f"【eteacher 売上反映】{result.output_path.name}",
        f"  テンプレ: {result.template_path.name}",
        f"  照合成功: {len(result.matched)}塾  内訳: {method_parts}",
        f"  eteacher 側で未マッチ: {len(result.unmatched_in_eteacher)}塾",
        f"  source 側で未マッチ: {len(result.unmatched_in_source)}塾",
    ]
    if result.unmatched_in_eteacher:
        lines.append("  先頭サンプル (eteacher 側):")
        for name in result.unmatched_in_eteacher[:detail_preview]:
            lines.append(f"    - {name}")
    if result.unmatched_in_source:
        lines.append("  先頭サンプル (source 側):")
        for name in result.unmatched_in_source[:detail_preview]:
            lines.append(f"    - {name}")
    return "\n".join(lines)


def write_unmatched_report(result: EteacherUpdateResult, report_path: str | Path) -> Path:
    """Write the full list of unmatched shops to a text file for manual review.

    目視で似た塾名を統一するときに使う。
    """
    report = Path(report_path)
    report.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        f"# eteacher 更新 未マッチ塾リスト",
        f"出力ファイル: {result.output_path}",
        f"テンプレ:     {result.template_path}",
        "",
        f"## eteacher 側にあるが source にない塾 ({len(result.unmatched_in_eteacher)}件)",
        "これらは eteacher の行は残るが、AO 列(売上)が空のままになります。",
        "source 側の塾名の表記揺れが原因なら手動で統一してください。",
        "",
    ]
    lines.extend(f"- {n}" for n in result.unmatched_in_eteacher)
    lines.append("")
    lines.append(f"## source 側にあるが eteacher にない塾 ({len(result.unmatched_in_source)}件)")
    lines.append("これらの売上は eteacher に載らないので、新規塾なら手動で行追加が必要です。")
    lines.append("")
    lines.extend(f"- {n}" for n in result.unmatched_in_source)
    report.write_text("\n".join(lines), encoding="utf-8")
    return report
