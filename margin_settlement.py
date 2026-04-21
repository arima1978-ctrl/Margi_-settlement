"""CLI entry point for margin settlement generation."""
from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import yaml

from src.new_shop_detector import detect_new_family_ids
from src.report_updater import append_family_ids_to_report, update_month_cell
from src.sales_master_sync import fetch_csv, sync_sales_management
from src.sheet_replacer import replace_sheet_data


def load_config(service: str) -> dict:
    config_path = Path(__file__).parent / "config" / f"{service}.yaml"
    with config_path.open(encoding="utf-8") as f:
        return yaml.safe_load(f)


def parse_month(month_str: str) -> datetime:
    """Parse YYYY-MM -> datetime(YYYY, MM, 1)."""
    year, month = month_str.split("-")
    return datetime(int(year), int(month), 1)


def parse_id_list(raw: str | None) -> list[int]:
    """Parse '101297,102676' -> [101297, 102676]. Empty string → []."""
    if not raw:
        return []
    out = []
    for item in raw.split(","):
        item = item.strip()
        if not item:
            continue
        out.append(int(item))
    return out


def run(service: str, source: str, template: str, output: str, month: str,
        skip_google_sheet: bool = False, add_shops: list[int] | None = None,
        sales_master_csv: str | None = None) -> None:
    print(f"[margin_settlement] service={service} month={month}")
    print(f"  source:   {source}")
    print(f"  template: {template}")
    print(f"  output:   {output}")

    cfg = load_config(service)
    month_date = parse_month(month)

    # 1. Copy template to output
    print("\n[1/5] Copying template...")
    shutil.copy2(template, output)

    # 2. Load source (data_only=True to resolve formulas to values)
    print("[2/5] Loading source .xlsm (resolving formulas to values)...")
    wb_src = openpyxl.load_workbook(source, data_only=True, keep_vba=False)

    # 3. Load output workbook (keep formulas)
    print("[3/5] Loading output workbook...")
    wb_out = openpyxl.load_workbook(output, data_only=False)

    # 4. Replace each configured sheet
    print("[4/5] Replacing sheets...")
    for mapping in cfg["sheet_replacements"]:
        src_name = mapping["source_sheet"]
        dst_name = mapping["dest_sheet"]
        data_start_row = mapping.get("data_start_row", mapping.get("header_rows", 1) + 1)
        helper_cols = mapping.get("helper_columns") or {}

        if src_name not in wb_src.sheetnames:
            print(f"  WARN: source sheet '{src_name}' not found, skipping")
            continue
        if dst_name not in wb_out.sheetnames:
            print(f"  WARN: dest sheet '{dst_name}' not found, skipping")
            continue

        src_ws = wb_src[src_name]
        dst_ws = wb_out[dst_name]
        aggregates = mapping.get("aggregates") or []
        rows = replace_sheet_data(src_ws, dst_ws, data_start_row, helper_cols, aggregates)
        print(f"  '{src_name}' -> '{dst_name}': {rows} rows")

    # 4b. Sync プログラミング営業管理 from Google Sheet CSV (optional)
    sms = cfg.get("sales_master_sync", {})
    if sms.get("enabled"):
        csv_source = sales_master_csv or sms.get("csv_url")
        if csv_source:
            print(f"\n[4b] Syncing '{sms['sheet']}' from CSV...")
            print(f"     source: {csv_source[:80]}{'...' if len(csv_source) > 80 else ''}")
            try:
                csv_rows = fetch_csv(csv_source)
                print(f"     CSV rows: {len(csv_rows)}")
                # Parse column_mapping: YAML may give string keys, ensure ints
                raw_map = sms["column_mapping"]
                col_map = {int(k): v for k, v in raw_map.items()}
                written = sync_sales_management(
                    ws=wb_out[sms["sheet"]],
                    csv_rows=csv_rows,
                    column_mapping=col_map,
                    csv_data_start_row=sms["csv_data_start_row"],
                    excel_data_start_row=sms["excel_data_start_row"],
                    excel_max_col_letter=sms.get("excel_max_col_letter", "AG"),
                )
                print(f"     Wrote {written} shops to '{sms['sheet']}'")
            except Exception as e:
                print(f"     WARN: Sales master sync failed: {e}")
                print(f"     Continuing without sync.")
        else:
            print("\n[4b] sales_master_sync.enabled=true but no csv_url/--sales-master given, skipping")

    # 5. Update 報告書!D1
    print("\n[5/5] Updating report metadata...")
    report_ws = wb_out[cfg["report_sheet"]]
    update_month_cell(report_ws, cfg["month_cell"], month_date)
    print(f"  {cfg['report_sheet']}!{cfg['month_cell']} = {month_date.date()}")

    # 6. New shop detection + manual add
    nsd = cfg.get("new_shop_detection", {})
    add_shops = add_shops or []
    if add_shops:
        print(f"\n[Extra] Adding user-specified shops: {add_shops}")
        rows_used = append_family_ids_to_report(
            report_ws, add_shops,
            data_start_row=nsd.get("report_data_start_row", 11),
        )
        print(f"  Appended {len(rows_used)} rows to 報告書: {rows_used}")

    if nsd.get("enabled"):
        print("\n[Extra] Detecting candidate new family IDs (active in ④ sheets but not in 報告書)...")
        new_ids = detect_new_family_ids(
            wb_src=wb_src,
            category_sheets=nsd["source_category_sheets"],
            dst_report_ws=report_ws,
            dst_report_id_column=nsd["report_family_id_column"],
            dst_report_start_row=nsd["report_data_start_row"],
        )
        if new_ids:
            print(f"  Found {len(new_ids)} candidate new family IDs (active in ④ sheets but not in 報告書):")
            print(f"    {new_ids}")
            auto_append = nsd.get("auto_append_to_report", False)
            if auto_append:
                rows_used = append_family_ids_to_report(
                    report_ws, new_ids,
                    data_start_row=nsd["report_data_start_row"],
                )
                print(f"  Appended to 報告書 rows: {rows_used}")
            else:
                print("  auto_append_to_report=false → 報告書への追加はスキップ（手動で確認してください）")

            if auto_append and not skip_google_sheet:
                try:
                    from src.google_sheet_reader import (
                        find_shop_rows_by_family_ids,
                        load_google_sheet_rows,
                    )
                    print("  Fetching shop info from Google Sheet...")
                    rows = load_google_sheet_rows(
                        nsd["google_sheet_id"], nsd["google_sheet_tab"]
                    )
                    matches = find_shop_rows_by_family_ids(rows, new_ids)
                    print(f"  Matched {len(matches)} / {len(new_ids)} shops in Google Sheet")
                    _append_to_sales_management(
                        wb_out[nsd["sales_management_sheet"]],
                        matches,
                        id_column=nsd["sales_management_id_column"],
                    )
                except Exception as e:
                    print(f"  WARN: Google Sheet fetch failed: {e}")
                    print("       Continuing without sales management population.")
        else:
            print("  No new family IDs detected.")

    # 7. Mark for full recalc when Excel opens the file
    wb_out.calculation.fullCalcOnLoad = True

    # 8. Save
    print("\nSaving...")
    wb_out.save(output)
    print(f"Done: {output}")


def _append_to_sales_management(ws, matches, id_column):
    """Stub — actual column mapping to be refined after inspecting Google Sheet."""
    from src.sheet_replacer import column_letter_to_index
    col_idx = column_letter_to_index(id_column)
    # Find first empty row in ID column
    row = 2
    while ws.cell(row=row, column=col_idx).value is not None:
        row += 1
    for fid, gs_row in matches.items():
        ws.cell(row=row, column=col_idx).value = fid
        # TODO: map Google Sheet columns to 営業管理 columns once structure confirmed
        row += 1


def main():
    parser = argparse.ArgumentParser(description="Generate monthly margin settlement")
    parser.add_argument("service", choices=["programming", "shogi", "bunri", "sokudoku"])
    parser.add_argument("--source", required=True, help="Source .xlsm (invoice data)")
    parser.add_argument("--template", required=True, help="Previous month's settlement")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    parser.add_argument("--month", required=True, help="Target month YYYY-MM")
    parser.add_argument("--skip-google-sheet", action="store_true",
                        help="Skip Google Sheet lookup for new shops")
    parser.add_argument("--add-shops", default="",
                        help="Comma-separated family IDs to append to 報告書 B/AL columns")
    parser.add_argument("--sales-master", default="",
                        help="URL or local path for プログラミング営業管理 CSV (overrides config)")
    args = parser.parse_args()

    try:
        run(args.service, args.source, args.template, args.output, args.month,
            skip_google_sheet=args.skip_google_sheet,
            add_shops=parse_id_list(args.add_shops),
            sales_master_csv=args.sales_master or None)
    except Exception as e:
        print(f"\nERROR: {e}", file=sys.stderr)
        raise


if __name__ == "__main__":
    main()
