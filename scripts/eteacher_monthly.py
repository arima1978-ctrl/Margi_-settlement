"""eteacher 売上管理表 月次自動生成.

Usage
-----
    # 自動: 今月の1日なら前月分を生成 (5/1 実行 → 2026-04 対象)
    python scripts/eteacher_monthly.py --notify

    # 明示
    python scripts/eteacher_monthly.py --month 2026-05 --notify

処理内容
----
1. 前月の eteacher .xlsx を探してテンプレとして複製
2. 当月用の 2 列 (入金日 / 売上額) を最右端に追加
3. source .xlsm の マージン計算用 から当月の 家族ID別売上 を計算
   source は target_month の 20日±5日 に作成された xlsm を優先
4. eteacher に無い新規家族IDを 行末に追加 (塾名/住所/TEL/代表者を自動補完)
5. 家族IDベースで当月列に売上を反映
6. Telegram 通知 (--notify)
"""
from __future__ import annotations

import argparse
import os
import re
import shutil
import sys
from copy import copy as copy_obj
from datetime import date
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

import openpyxl
from openpyxl.utils import get_column_letter

from src.eteacher_updater import (
    _build_hogosha_info_lookup,
    _build_rakuraku_rep_lookup,
    compute_shops_with_sales,
    convert_xls_to_xlsx,
)
from src.notifier import load_dotenv, send_telegram

load_dotenv(REPO_ROOT / ".env", override=True)

BASE_DIR = Path(os.environ.get("MARGIN_BASE_DIR") or r"Y:\_★20170701作業用\【エデュプラス請求書】")
SOURCE_BACKUP_DIR = BASE_DIR / "【業者請求書】エクセルbackup"
ETEACHER_DIR = BASE_DIR.parent / "9三浦"

DATA_START_ROW = 8
COL_ACTIVE = 1     # A = 有効フラグ
COL_FAMILY_ID = 4  # D
COL_NAME = 5       # E
COL_REP = 7        # G 代表者
COL_ADDR = 8       # H 住所
COL_TEL = 9        # I 電話番号


# ---- month / file helpers ----

def prev_month(d: date) -> date:
    if d.month == 1:
        return date(d.year - 1, 12, 1)
    return date(d.year, d.month - 1, 1)


def find_source_xlsm(target_month: date, day_target: int = 20, tol: int = 5) -> Path:
    """target_month の day_target ± tol 日の送信分 folder → .xlsm."""
    pattern = f"{target_month.year}年{target_month.month}月*日送信分"
    candidates: list[tuple[int, Path]] = []
    for folder in SOURCE_BACKUP_DIR.glob(pattern):
        m = re.search(rf"{target_month.month}月(\d+)日送信分", folder.name)
        if not m:
            continue
        day = int(m.group(1))
        if abs(day - day_target) > tol:
            continue
        candidates.append((abs(day - day_target), folder))
    if not candidates:
        # fallback: 全ての当月送信分から、20日に近い方
        for folder in SOURCE_BACKUP_DIR.glob(pattern):
            m = re.search(rf"{target_month.month}月(\d+)日送信分", folder.name)
            if m:
                day = int(m.group(1))
                candidates.append((abs(day - day_target), folder))
    if not candidates:
        raise FileNotFoundError(
            f"送信分 folder not found for {target_month.year}年{target_month.month}月"
        )
    candidates.sort()
    src_folder = candidates[0][1]
    for f in src_folder.glob("*.xlsm"):
        if f.name.startswith("~$") or "入金チェック" in f.name:
            continue
        return f
    raise FileNotFoundError(f".xlsm not found in {src_folder}")


def find_prev_eteacher(target_month: date) -> Path:
    prev = prev_month(target_month)
    for ext in [".xlsx", ".xls"]:
        p = ETEACHER_DIR / f"eteacher売上管理表{prev.year}年{prev.month}月{ext}"
        if p.exists():
            return p
    raise FileNotFoundError(
        f"前月の eteacher が見つからない: eteacher売上管理表{prev.year}年{prev.month}月.xlsx"
    )


def output_path(target_month: date) -> Path:
    return ETEACHER_DIR / f"eteacher売上管理表{target_month.year}年{target_month.month}月.xlsx"


# ---- sheet manipulation ----

def _last_sales_col(ws) -> int:
    """row 7 で最後の '売上額' ヘッダの列番号."""
    last = 0
    max_col = max(ws.max_column, 60)
    for c in range(1, max_col + 1):
        if ws.cell(7, c).value == "売上額":
            last = c
    return last


def _last_data_row(ws) -> int:
    last = DATA_START_ROW - 1
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if ws.cell(r, COL_NAME).value:
            last = r
    return last


def add_monthly_columns(ws, target_month: date) -> tuple[int, int]:
    """最後の '売上額' 列の右に 2 列 (入金日/売上額) を挿入して返す.

    挿入方式のため、AP-AS にある参照列 (住所/TEL/代表者/照合方法) は
    AR-AV に自動的にシフトされる (openpyxl insert_cols)。
    """
    last = _last_sales_col(ws)
    if last == 0:
        # fallback: 右端に追記
        last = ws.max_column

    # last+1 の位置に 2 列挿入 (既存の右側データを 2 列右にシフト)
    ws.insert_cols(idx=last + 1, amount=2)

    date_col = last + 1
    sales_col = last + 2

    # r7 ヘッダ
    ws.cell(row=7, column=date_col).value = "入金日"
    ws.cell(row=7, column=sales_col).value = "売上額"

    # r6 月ラベル (表示用)
    month_label = f"{target_month.year}年{target_month.month}月"
    if not ws.cell(row=6, column=date_col).value:
        ws.cell(row=6, column=date_col).value = month_label

    # r1/r2 合計
    last_row = _last_data_row(ws)
    sales_letter = get_column_letter(sales_col)
    ws.cell(row=1, column=sales_col).value = f"=SUM({sales_letter}8:{sales_letter}{last_row + 50})"
    ws.cell(row=2, column=sales_col).value = f"={sales_letter}1*0.18"
    ws.cell(row=1, column=sales_col).number_format = "#,##0"
    ws.cell(row=2, column=sales_col).number_format = "#,##0"

    # 前月の sales 列の書式を新列にコピー (罫線/フォント)
    # 元の last は insert_cols で位置が変わっていないので、そのまま last を参照する
    _copy_column_style(ws, last, sales_col)
    _copy_column_style(ws, last - 1, date_col)

    return date_col, sales_col


def _copy_column_style(ws, src_col: int, dst_col: int, max_row: int = 2000) -> None:
    """src_col の各行のセルスタイルを dst_col にコピー."""
    src_letter = get_column_letter(src_col)
    dst_letter = get_column_letter(dst_col)
    # 列幅
    src_width = ws.column_dimensions[src_letter].width
    if src_width:
        ws.column_dimensions[dst_letter].width = src_width
    # 行ごとのスタイル (最大 max_row 行)
    limit = min(max_row, ws.max_row)
    for r in range(1, limit + 1):
        s = ws.cell(row=r, column=src_col)
        d = ws.cell(row=r, column=dst_col)
        if s.has_style:
            d.font = copy_obj(s.font)
            d.fill = copy_obj(s.fill)
            d.border = copy_obj(s.border)
            d.alignment = copy_obj(s.alignment)
            if dst_col % 2 == 0:
                # 売上列は数値書式
                d.number_format = "#,##0"
            else:
                d.number_format = s.number_format


def fill_monthly_sales(ws, shops, sales_col: int) -> tuple[int, int]:
    """sales_col に売上を書き込む. (matched_count, total_amount)."""
    by_fid = {s.family_id: s.sales for s in shops}
    matched = 0
    total = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        fid_raw = ws.cell(r, COL_FAMILY_ID).value
        if fid_raw is None or fid_raw == "":
            continue
        try:
            fid = int(fid_raw)
        except (TypeError, ValueError):
            continue
        if fid == 0:
            continue
        sales = by_fid.get(fid, 0)
        cell = ws.cell(row=r, column=sales_col)
        if sales != 0:
            cell.value = sales
            cell.number_format = "#,##0"
            matched += 1
            total += sales
        else:
            cell.value = None
    return matched, total


def append_new_shops(ws, shops, existing_fids: set[int]) -> list[dict]:
    """source にあって eteacher に無い 家族ID を 行末に追加. 追加リストを返す."""
    start_row = _last_data_row(ws) + 1
    added: list[dict] = []
    cursor = start_row
    for shop in shops:
        if shop.family_id in existing_fids or shop.sales == 0:
            continue
        ws.cell(row=cursor, column=COL_ACTIVE).value = 1
        ws.cell(row=cursor, column=COL_FAMILY_ID).value = shop.family_id
        ws.cell(row=cursor, column=COL_NAME).value = shop.name or ""
        if shop.rep:
            ws.cell(row=cursor, column=COL_REP).value = shop.rep
        if shop.addr:
            ws.cell(row=cursor, column=COL_ADDR).value = shop.addr
        if shop.tel:
            ws.cell(row=cursor, column=COL_TEL).value = shop.tel
        added.append({
            "row": cursor,
            "family_id": shop.family_id,
            "name": shop.name or "",
            "sales": shop.sales,
        })
        cursor += 1
    return added


def collect_existing_family_ids(ws) -> set[int]:
    out = set()
    for r in range(DATA_START_ROW, ws.max_row + 1):
        v = ws.cell(r, COL_FAMILY_ID).value
        if v is None or v == "":
            continue
        try:
            fid = int(v)
            if fid > 0:
                out.add(fid)
        except (TypeError, ValueError):
            continue
    return out


# ---- main ----

def _resolve_target_month(month_arg: str | None) -> date:
    if month_arg:
        y, m = month_arg.split("-")
        return date(int(y), int(m), 1)
    today = date.today()
    return prev_month(date(today.year, today.month, 1))


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--month", help="対象月 YYYY-MM (省略時は前月)")
    parser.add_argument("--source", help="source .xlsm を明示")
    parser.add_argument("--template", help="前月 eteacher テンプレを明示")
    parser.add_argument("--output", help="出力 .xlsx パスを明示")
    parser.add_argument("--notify", action="store_true", help="Telegram 通知")
    parser.add_argument("--no-backup", action="store_true", help=".bak を作らない")
    args = parser.parse_args()

    target = _resolve_target_month(args.month)
    print(f"対象月: {target.year}-{target.month:02d}")

    try:
        source = Path(args.source) if args.source else find_source_xlsm(target)
        template = Path(args.template) if args.template else find_prev_eteacher(target)
    except FileNotFoundError as exc:
        msg = f"【eteacher 月次】{target.year}-{target.month:02d} ファイル未検出: {exc}"
        print(msg, file=sys.stderr)
        if args.notify:
            send_telegram(msg)
        return 2

    out = Path(args.output) if args.output else output_path(target)
    out.parent.mkdir(parents=True, exist_ok=True)

    print(f"source:   {source}")
    print(f"template: {template}")
    print(f"output:   {out}")
    print()

    # Step 1: copy template → output (.xls なら .xlsx に変換)
    print("[1/5] 前月テンプレを複製")
    if template.suffix.lower() == ".xls":
        print(f"     .xls → .xlsx 変換中: {template.name}")
        convert_xls_to_xlsx(template, out)
    else:
        shutil.copy2(template, out)

    # Step 2: compute shops with sales from source
    print("[2/5] source から 家族ID別売上を計算")
    shops = compute_shops_with_sales(source)
    print(f"     source塾数 (売上>0含む): {len(shops)}")

    # Step 3: add 2 new columns
    print("[3/5] 当月の入金日/売上額列を追加")
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    date_col, sales_col = add_monthly_columns(ws, target)
    print(f"     入金日列 = {get_column_letter(date_col)}, 売上額列 = {get_column_letter(sales_col)}")

    # Step 4: append new family IDs as rows
    print("[4/5] 新規家族ID を行末に追加")
    existing_fids = collect_existing_family_ids(ws)
    added = append_new_shops(ws, shops, existing_fids)
    print(f"     新規追加: {len(added)} 塾")
    for info in added:
        print(f"       r{info['row']} ID={info['family_id']} {info['name']!r} 売上={info['sales']:,}")

    # Step 5: fill sales column
    print("[5/5] 当月売上を反映")
    matched, total = fill_monthly_sales(ws, shops, sales_col)
    print(f"     反映: {matched} 塾, 売上合計={total:,} 円")

    # Save
    wb.save(out)
    wb.close()

    # Summary (Telegram 本文は新規追加塾を中心に)
    summary_lines = [
        f"【eteacher 月次生成】{target.year}年{target.month}月",
        f"売上合計: {total:,} 円 ({matched} 塾)",
        "",
    ]
    if added:
        summary_lines.append(f"■ 新規追加塾 ({len(added)} 件):")
        for info in added:
            name = info["name"] or "(塾名なし)"
            sales = info["sales"]
            sales_txt = f"売上 {sales:,} 円" if sales else "売上 0 円"
            summary_lines.append(f"  ・ID={info['family_id']} {name} ({sales_txt})")
    else:
        summary_lines.append("■ 新規追加塾: なし")
    summary_lines.append("")
    summary_lines.append(f"source: {source.name}")
    summary_lines.append(f"output: {out.name}")
    summary = "\n".join(summary_lines)
    print()
    print(summary)

    if args.notify:
        ok = send_telegram(summary)
        print(f"\nTelegram: {'送信OK' if ok else '送信失敗'}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
