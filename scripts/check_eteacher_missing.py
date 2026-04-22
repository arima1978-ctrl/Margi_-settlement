"""eteacher.xlsx に反映されていない source 側の売上を洗い出す診断ツール.

Usage
-----
    python scripts/check_eteacher_missing.py \
      --target "C:\\...\\eteacher売上管理表2026年4月.xlsx" \
      --source "Y:\\...\\2026年4月17日送信.xlsm"
      [--report "C:\\...\\missing.md"]

3 種類の問題を検出する:
  ① source に売上 (!=0) があるが eteacher の D 列に家族ID として載っていない
  ② eteacher にあるが AO の数値が source と食い違う
  ③ eteacher の D 列 家族ID が空の行 (塾名で候補があれば表示)
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

import openpyxl

from src.eteacher_updater import compute_shops_with_sales, normalize_shop_name


DATA_START_ROW = 8
COL_FAMILY_ID = 4   # D
COL_NAME = 5        # E
COL_SALES = 41      # AO


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--target", required=True)
    parser.add_argument("--source", required=True)
    parser.add_argument("--report", help="レポート出力パス (省略時は target の隣)")
    args = parser.parse_args()

    target = Path(args.target)
    source = Path(args.source)
    report = Path(args.report) if args.report else target.with_suffix(".missing.md")

    print(f"[1/3] source 売上読込: {source.name}")
    shops = compute_shops_with_sales(source)
    source_sales = {s.family_id: s for s in shops if s.sales != 0}
    print(f"     売上!=0 の塾: {len(source_sales)}")

    print(f"[2/3] eteacher 走査: {target.name}")
    wb = openpyxl.load_workbook(target, data_only=False)
    ws = wb.active
    eteacher_rows: dict[int, tuple[int, str, object]] = {}
    no_fid_rows: list[tuple[int, str]] = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        fid_raw = ws.cell(row, COL_FAMILY_ID).value
        name = ws.cell(row, COL_NAME).value
        ao = ws.cell(row, COL_SALES).value
        if not isinstance(name, str) or not name.strip():
            continue
        if fid_raw is None or fid_raw == "":
            no_fid_rows.append((row, name.strip()))
            continue
        try:
            fid = int(fid_raw)
        except (TypeError, ValueError):
            no_fid_rows.append((row, name.strip()))
            continue
        eteacher_rows[fid] = (row, name.strip(), ao)
    wb.close()

    missing_from_eteacher = [s for fid, s in source_sales.items() if fid not in eteacher_rows]
    mismatches = []
    for fid, (row, name, ao) in eteacher_rows.items():
        s = source_sales.get(fid)
        if s is None:
            continue
        actual = int(ao) if isinstance(ao, (int, float)) else None
        if actual != s.sales:
            mismatches.append((row, fid, name, s.sales, ao))

    # Source-by-canonical-name for suggesting matches for no-fid rows
    source_by_canon = {}
    for s in shops:
        c = normalize_shop_name(s.name)
        if c:
            source_by_canon.setdefault(c, s)

    buf: list[str] = [
        f"# 売上取りこぼし診断",
        f"対象: {target}",
        f"源泉: {source}",
        "",
        f"- source (売上!=0): {len(source_sales)} 塾",
        f"- eteacher (家族ID入り): {len(eteacher_rows)} 塾",
        "",
        f"## ① source に売上あるが eteacher 未登録 — {len(missing_from_eteacher)} 件",
        "eteacher に新規行として追加してから refresh 実行する必要あり。",
    ]
    for s in sorted(missing_from_eteacher, key=lambda x: -abs(x.sales)):
        buf.append(f"  - ID={s.family_id:>6}  塾名={s.name!r:<40s}  売上={s.sales:>8d}  TEL={s.tel!r}")

    buf.append("")
    buf.append(f"## ② AO 値が source と不一致 — {len(mismatches)} 件")
    if mismatches:
        buf.append("(refresh 済みなら 0 件。残っていれば refresh_eteacher を再実行)")
        for row, fid, name, expected, actual in mismatches[:100]:
            buf.append(f"  - r{row:<4} ID={fid:>6}  {name!r:<30s}  期待={expected:>8d}  実際={actual!r}")
        if len(mismatches) > 100:
            buf.append(f"  ... 他 {len(mismatches) - 100} 件")
    else:
        buf.append("問題なし ✓")

    buf.append("")
    buf.append(f"## ③ eteacher の D 列 家族ID 未入力 — {len(no_fid_rows)} 件")
    for row, name in no_fid_rows:
        c = normalize_shop_name(name)
        hit = source_by_canon.get(c)
        if hit and hit.sales != 0:
            buf.append(f"  - r{row:<4}  塾名={name!r}  → 候補: ID={hit.family_id} source塾名={hit.name!r} 売上={hit.sales}")
        else:
            buf.append(f"  - r{row:<4}  塾名={name!r}")

    report.write_text("\n".join(buf), encoding="utf-8")
    print(f"[3/3] レポート: {report}")

    print()
    print("【概要】")
    print(f"  ① source にあるが eteacher 未登録: {len(missing_from_eteacher)} 件")
    print(f"  ② AO 値が source と不一致:         {len(mismatches)} 件")
    print(f"  ③ 家族ID 未入力:                   {len(no_fid_rows)} 件")
    if missing_from_eteacher:
        total = sum(s.sales for s in missing_from_eteacher)
        print(f"  取りこぼし額合計: {total:,} 円")

    return 0 if not missing_from_eteacher and not mismatches else 1


if __name__ == "__main__":
    sys.exit(main())
