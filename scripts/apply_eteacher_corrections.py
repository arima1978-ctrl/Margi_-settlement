"""診断 xlsx の J列 (修正家族ID) を読み取り、eteacher .xlsx の該当行 D列
(家族ID) に反映する。

Usage
-----
    python scripts/apply_eteacher_corrections.py \
      --corrections "C:\\tmp\\margin_inspect\\id_name_mismatch_full.xlsx" \
      --target      "Y:\\_★20170701作業用\\9三浦\\eteacher売上管理表2026年4月.xlsx"

workflow:
  1. 診断 xlsx を Excel で開き、J列に正しい家族IDを記入 (複数行 OK)
  2. 保存して Excel を閉じる
  3. 本ツール実行 → eteacher の D列 を上書き
  4. refresh_eteacher.py で 売上再反映、check_eteacher_missing.py で最終確認

Columns expected in the corrections xlsx:
  A: 行(eteacher の行番号)
  B: 現 家族ID (参考)
  J: 修正家族ID (空なら無視、数字ならそれで上書き)
"""
from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

import openpyxl

COL_CORRECTIONS_ROW = 1   # A = eteacher 行番号
COL_CORRECTIONS_OLD = 2   # B = 旧家族ID (参考)
COL_CORRECTIONS_NEW = 10  # J = 修正家族ID
COL_ETEACHER_FID = 4      # D = 家族ID


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--corrections", required=True,
                        help="修正家族IDを書いた診断 xlsx のパス")
    parser.add_argument("--target", required=True,
                        help="反映先の eteacher .xlsx のパス")
    parser.add_argument("--no-backup", action="store_true", help=".bak を作らない")
    args = parser.parse_args()

    corrections_path = Path(args.corrections)
    target_path = Path(args.target)

    if not args.no_backup:
        bak = target_path.with_suffix(target_path.suffix + ".bak")
        shutil.copy2(target_path, bak)
        print(f"[backup] {bak}")

    # 1. 診断 xlsx から J列 に値があるもののみ収集
    print(f"[1/2] 修正リスト読込: {corrections_path}")
    wb_c = openpyxl.load_workbook(corrections_path, data_only=True, read_only=True)
    ws_c = wb_c.active
    corrections: list[tuple[int, int, int]] = []   # (eteacher_row, old_fid, new_fid)
    for row in ws_c.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        et_row = row[COL_CORRECTIONS_ROW - 1]
        old_fid = row[COL_CORRECTIONS_OLD - 1]
        new_fid = row[COL_CORRECTIONS_NEW - 1] if len(row) >= COL_CORRECTIONS_NEW else None
        if et_row is None or new_fid is None or new_fid == "":
            continue
        try:
            et_row_i = int(et_row)
            new_fid_i = int(new_fid)
            old_fid_i = int(old_fid) if old_fid else None
        except (TypeError, ValueError):
            continue
        corrections.append((et_row_i, old_fid_i, new_fid_i))
    wb_c.close()
    print(f"     修正対象: {len(corrections)} 件")

    if not corrections:
        print("  J列に修正家族IDが入っていません。終了します。")
        return 0

    # 2. eteacher に反映
    print(f"[2/2] eteacher へ反映: {target_path}")
    wb = openpyxl.load_workbook(target_path)
    ws = wb.active

    applied = 0
    skipped = 0
    for et_row, old_fid, new_fid in corrections:
        current = ws.cell(row=et_row, column=COL_ETEACHER_FID).value
        try:
            current_i = int(current) if current is not None and current != "" else None
        except (TypeError, ValueError):
            current_i = None
        if old_fid is not None and current_i != old_fid:
            print(f"  WARN r{et_row}: 現在値={current_i} が予期した旧ID={old_fid} と異なる → スキップ")
            skipped += 1
            continue
        ws.cell(row=et_row, column=COL_ETEACHER_FID).value = new_fid
        applied += 1
        print(f"  r{et_row}: D列 {current_i} → {new_fid}")

    wb.save(target_path)
    wb.close()

    print()
    print(f"反映完了: {applied} 件  /  スキップ: {skipped} 件")
    if applied:
        print()
        print("次のステップ:")
        print("  1. refresh_eteacher.py で 売上を再反映")
        print("  2. check_eteacher_missing.py で最終確認")
    return 0


if __name__ == "__main__":
    sys.exit(main())
