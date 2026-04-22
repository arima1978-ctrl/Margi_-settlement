"""Verify bunri: regenerate March 2026 settlement, diff against existing file."""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

from margin_settlement import run

FEB_TEMPLATE = r"Y:\_★20170701作業用\【エデュプラス請求書】\文理ヴィクトリー清算書\文理ヴィクトリー売上管理表_202602月分.xlsx"
SOURCE_XLSM  = r"Y:\_★20170701作業用\【エデュプラス請求書】\【業者請求書】エクセルbackup\2026年2月17日送信分\2026年2月17日送信.xlsm"
EXPECTED_MAR = r"Y:\_★20170701作業用\【エデュプラス請求書】\文理ヴィクトリー清算書\文理ヴィクトリー売上管理表_202603月分.xlsx"

OUTPUT_DIR = Path(r"C:\tmp\margin_explore")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
GENERATED = str(OUTPUT_DIR / "generated_bunri_202603月分.xlsx")

SHEETS_TO_CHECK = [
    "報告書",
    "保護者情報DL貼付⑩AKへ",
    "文理ビクトリー_基本料金",
    "文理_生徒",
    "文理V_オンライン個別",
    "文理ヴィクトリー有料講座",
    "文理おまかせコース",
    "文理ヴィクトリー初期費用",
]


def normalize(v):
    if isinstance(v, ArrayFormula):
        return ("AF", v.ref, v.text)
    return v


def diff_sheet(gen_ws, exp_ws, sheet_name, max_reports=30):
    diffs = []
    max_row = max(gen_ws.max_row, exp_ws.max_row)
    max_col = max(gen_ws.max_column, exp_ws.max_column)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            gv = normalize(gen_ws.cell(row=r, column=c).value)
            ev = normalize(exp_ws.cell(row=r, column=c).value)
            if gv != ev:
                addr = gen_ws.cell(row=r, column=c).coordinate
                diffs.append((addr, gv, ev))
    print(f"  [{sheet_name}] {len(diffs)} diffs")
    for addr, gv, ev in diffs[:max_reports]:
        print(f"    {addr}: GEN={gv!r}  EXP={ev!r}")
    if len(diffs) > max_reports:
        print(f"    ... and {len(diffs) - max_reports} more")
    return len(diffs)


def main():
    print("=" * 80)
    print("Regenerating bunri March 2026 settlement from Feb template...")
    print("=" * 80)
    run(
        service="bunri",
        source=SOURCE_XLSM,
        template=FEB_TEMPLATE,
        output=GENERATED,
        month="2026-03",
        skip_google_sheet=True,
    )

    print("\n" + "=" * 80)
    print("Comparing generated vs existing March file...")
    print("=" * 80)
    wb_gen = openpyxl.load_workbook(GENERATED, data_only=False)
    wb_exp = openpyxl.load_workbook(EXPECTED_MAR, data_only=False)

    total = 0
    for sheet in SHEETS_TO_CHECK:
        if sheet not in wb_gen.sheetnames or sheet not in wb_exp.sheetnames:
            print(f"  [{sheet}] missing in one of the files, skipping")
            continue
        total += diff_sheet(wb_gen[sheet], wb_exp[sheet], sheet)
    print(f"\nTotal diffs across checked sheets: {total}")


if __name__ == "__main__":
    main()
