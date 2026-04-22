"""既存の eteacher.xlsx に対し、家族ID(D列)から売上を引き当てて AO/参照列を
再書き込みする CLI。

ユーザーが D列に手入力で 家族ID を埋めた後の 「売上再反映」 に使う。

Usage
-----
    python scripts/refresh_eteacher.py \
      --target  "C:\\tmp\\margin_inspect\\eteacher_test\\eteacher売上管理表2026年4月.xlsx" \
      --source  "Y:\\...\\2026年4月17日送信.xlsm"

既存ファイルは書き換え前に ``.bak`` にコピーされる (--no-backup で抑止)。
"""
from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

import openpyxl

from src.eteacher_updater import compute_shops_with_sales
from src.notifier import load_dotenv

load_dotenv(REPO_ROOT / ".env", override=True)


# 4月ファイルの列構成 (家族ID挿入済 = 1列シフト後)
COL_FAMILY_ID = 4   # D = 家族ID (ユーザー入力済)
COL_NAME      = 5   # E = 塾名
COL_SALES     = 41  # AO = 当月売上
COL_REF_ADDR  = 42  # AP
COL_REF_TEL   = 43  # AQ
COL_REF_REP   = 44  # AR
COL_METHOD    = 45  # AS = 照合方法

DATA_START_ROW = 8


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--target", required=True, help="更新する eteacher .xlsx のパス")
    parser.add_argument("--source", required=True, help="売上データを持つ処理済 .xlsm のパス")
    parser.add_argument("--no-backup", action="store_true",
                        help=".xlsx.bak を作らない")
    args = parser.parse_args()

    target = Path(args.target)
    source = Path(args.source)
    if not target.exists():
        print(f"ERROR: target not found: {target}", file=sys.stderr)
        return 2
    if not source.exists():
        print(f"ERROR: source not found: {source}", file=sys.stderr)
        return 2

    if not args.no_backup:
        bak = target.with_suffix(target.suffix + ".bak")
        shutil.copy2(target, bak)
        print(f"[backup] {bak}")

    # --- 1) source から 家族ID -> ShopSales dict を構築
    print(f"[1/3] 売上データ読込: {source}")
    shops = compute_shops_with_sales(source)
    by_fid = {s.family_id: s for s in shops}
    print(f"     {len(by_fid)} 塾を読み込み")

    # --- 2) target を開いて 家族ID で引き当て、AO/AP/AQ/AR/AS を書き込み
    print(f"[2/3] eteacher 更新中: {target}")
    wb = openpyxl.load_workbook(target)
    ws = wb.active

    written = 0
    no_fid = 0
    no_match = 0
    sales_zero = 0

    for row in range(DATA_START_ROW, ws.max_row + 1):
        fid_raw = ws.cell(row=row, column=COL_FAMILY_ID).value
        if fid_raw is None or fid_raw == "":
            no_fid += 1
            continue
        try:
            fid = int(fid_raw)
        except (TypeError, ValueError):
            no_fid += 1
            continue

        shop = by_fid.get(fid)
        if shop is None:
            no_match += 1
            # source 側に存在しない家族ID → AO/参照列は触らない (既存値を保持)
            continue

        # 売上: 0 の場合は cell を空にする (前の値が残っていると誤解を招く)
        if shop.sales != 0:
            ws.cell(row=row, column=COL_SALES).value = shop.sales
        else:
            ws.cell(row=row, column=COL_SALES).value = None
            sales_zero += 1

        # 参照列を再書き込み (family_id 基準で確実なデータ)
        ws.cell(row=row, column=COL_REF_ADDR).value = shop.addr or None
        ws.cell(row=row, column=COL_REF_TEL).value = shop.tel or None
        ws.cell(row=row, column=COL_REF_REP).value = shop.rep or None
        ws.cell(row=row, column=COL_METHOD).value = "id"
        written += 1

    wb.save(target)
    wb.close()

    print(f"[3/3] 完了: {target}")
    print()
    print(f"  家族ID で売上反映: {written} 塾")
    if sales_zero:
        print(f"    うち 売上=0 (AO 空欄): {sales_zero} 塾")
    if no_fid:
        print(f"  D列に 家族ID が入っていない行: {no_fid} 件 (AO 更新せず)")
    if no_match:
        print(f"  D列に家族IDあるが source に存在しない: {no_match} 件 (AO 更新せず)")

    return 0


if __name__ == "__main__":
    sys.exit(main())
