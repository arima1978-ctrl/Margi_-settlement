"""eteacher r424 D列を 101082 (神戸松陰塾 泉台校) に直す一度きりの修正."""
import openpyxl
from pathlib import Path

p = Path(r"Y:\_★20170701作業用\9三浦\eteacher売上管理表2026年4月.xlsx")
wb = openpyxl.load_workbook(p)
ws = wb.active
before = ws.cell(row=424, column=4).value
ws.cell(row=424, column=4).value = 101082
wb.save(p)
wb.close()
print(f"r424 D列: {before} -> 101082 に更新完了")
